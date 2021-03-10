import openpyxl
import matplotlib.pyplot as plt

wb=openpyxl.load_workbook('Reservoir_ex1.xlsx',data_only=True)
sheet=wb.active

max_col=sheet.max_column
max_row=sheet.max_row

x=[]
y=[]
for i in range(3,max_row-1):
    cell_obj=sheet.cell(row=i,column=1)
    y.append(cell_obj.value)

for j in range(3,max_row-1):
    cell_obj=sheet.cell(row=j,column=10)
    x.append(cell_obj.value)

plt.plot(x,y)
plt.title('Pres. Vs Bo')
plt.show()
